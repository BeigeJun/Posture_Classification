import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset
import openpyxl

Basic_path = 'C:/Users/wns20/PycharmProjects/SMART_CCTV/'
Exel_file_path = Basic_path + 'Find_Parameters/Parameters_Exel/ParametersAndResults.xlsx'
csv_file_path = Basic_path + '05.MLP_With_Angle/Data/Angle_data.csv'
data = pd.read_csv(csv_file_path)


#CSV파일 -> DataSet
X = data.drop('label', axis=1).values
y = data['label'].values

label_encoder = LabelEncoder()
y = label_encoder.fit_transform(y)
# for i in range(len(y)):
#     print(y[i])

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

X_train_tensor = torch.tensor(X_train, dtype=torch.float32)
y_train_tensor = torch.tensor(y_train, dtype=torch.long)
X_test_tensor = torch.tensor(X_test, dtype=torch.float32)
y_test_tensor = torch.tensor(y_test, dtype=torch.long)

train_dataset = TensorDataset(X_train_tensor, y_train_tensor)
test_dataset = TensorDataset(X_test_tensor, y_test_tensor)
train_loader = DataLoader(train_dataset, batch_size=32, shuffle=True)
test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False)

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
print(f'Using device: {device}')


class MLP(nn.Module):
    def __init__(self, input_size, f1_num, f2_num, f3_num, f4_num, f5_num, f6_num, d1, d2, d3, d4, d5, num_classes):
        super(MLP, self).__init__()
        self.relu = nn.ReLU()
        self.softmax = nn.Softmax(dim=0)
        self.fc1 = nn.Linear(input_size, f1_num)
        self.fc2 = nn.Linear(f1_num, f2_num)
        self.fc3 = nn.Linear(f2_num, f3_num)
        self.fc4 = nn.Linear(f3_num, f4_num)
        self.fc5 = nn.Linear(f4_num, f5_num)
        self.fc6 = nn.Linear(f5_num, f6_num)
        self.fc7 = nn.Linear(f6_num, num_classes)
        self.dropout1 = nn.Dropout(p=d1)
        self.dropout2 = nn.Dropout(p=d2)
        self.dropout3 = nn.Dropout(p=d3)
        self.dropout4 = nn.Dropout(p=d4)
        self.dropout5 = nn.Dropout(p=d5)

    def forward(self, x):
        out = self.dropout1(x)
        out = self.fc1(out)
        out = self.relu(out)
        out = self.dropout2(out)
        out = self.fc2(out)
        out = self.relu(out)
        out = self.dropout3(out)
        out = self.fc3(out)
        out = self.relu(out)
        out = self.dropout4(out)
        out = self.fc4(out)
        out = self.relu(out)
        out = self.dropout3(out)
        out = self.fc5(out)
        out = self.relu(out)
        out = self.dropout2(out)
        out = self.fc6(out)
        out = self.relu(out)
        out = self.dropout1(out)
        out = self.fc7(out)
        return out


def initialize_weights(model):
    for layer in model.modules():
        if isinstance(layer, nn.Linear):
            nn.init.xavier_uniform_(layer.weight)
            if layer.bias is not None:
                nn.init.zeros_(layer.bias)


input_size = X_train.shape[1]
num_classes = len(label_encoder.classes_)
# Exel File -> Parameters
Exel_File = openpyxl.load_workbook(Exel_file_path)
Sheet_Read_Data = Exel_File['Parameters']
Sheet_Save_Data = Exel_File['Results']
for index, (input_num, h1, h2, h3, h4, h5, h6, output_num, d1, d2, d3, d4, d5, batch, optimizer, lr, epochs, name) in enumerate(Sheet_Read_Data.iter_rows(values_only=True)):

    if index == 0:
        continue
    model = MLP(input_size, int(h1), int(h2), int(h3), int(h4), int(h5), int(h6), float(d1), float(d2), float(d3), float(d4), float(d5), int(output_num)).to(device)
    initialize_weights(model)
    criterion = nn.CrossEntropyLoss()
    opt = optimizer == 'Adam'
    if opt:
        optimizer = optim.Adam(model.parameters(), lr=lr)
    else:
        optimizer = optim.SGD(model.parameters(), lr=lr)

    for epoch in range(int(epochs)):
        model.train()
        for i, (inputs, labels) in enumerate(train_loader):
            inputs, labels = inputs.to(device), labels.to(device)

            outputs = model(inputs)
            loss = criterion(outputs, labels)

            optimizer.zero_grad()
            loss.backward()
            optimizer.step()

        print(f'Model[{index}] Epoch [{epoch+1}/{epochs}], Loss: {loss.item():.10}')

    model_path = 'C:/Users/wns20/PycharmProjects/SMART_CCTV/05.MLP_With_Angle/5_1/' + str(name) + '.pth'
    torch.save(model.state_dict(), model_path)
    print(f"Model saved to {model_path}")

    model.eval()
    with torch.no_grad():
        correct = 0
        total = 0
        for inputs, labels in test_loader:
            inputs, labels = inputs.to(device), labels.to(device)

            outputs = model(inputs)
            _, predicted = torch.max(outputs.data, 1)
            total += labels.size(0)
            correct += (predicted == labels).sum().item()

        accuracy = correct / total
        print(f'Model Accuracy: {accuracy * 100:.2f}%')
        last_row = Sheet_Save_Data.max_row + 1
        Sheet_Save_Data.cell(row=last_row, column=1).value = name
        Sheet_Save_Data.cell(row=last_row, column=2).value = accuracy * 100
        Exel_File.save(Exel_file_path)